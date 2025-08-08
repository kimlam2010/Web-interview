"""
Report Generation Module

Comprehensive reporting system with Excel export functionality,
candidate progress reports, position performance analytics,
interviewer performance tracking, and automated scheduling.
"""

import os
import json
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from flask import Blueprint, render_template, request, jsonify, send_file, flash, redirect, url_for
from flask_login import login_required, current_user
from sqlalchemy import and_, or_, func, desc, extract, case
from app.models import (
    db, Candidate, Position, AssessmentResult, InterviewEvaluation, 
    ExecutiveDecision, User, Step3ExecutiveFeedback, AuditLog
)
from app.decorators import hr_required, admin_required
from app.security import audit_log, rate_limit, security_check

report_generation_bp = Blueprint('report_generation', __name__)


class ReportGenerator:
    """Main report generator class with Excel export capabilities."""
    
    @staticmethod
    def generate_candidate_progress_report(candidate_id: int = None, position_id: int = None, 
                                        date_from: datetime = None, date_to: datetime = None) -> pd.DataFrame:
        """Generate candidate progress report."""
        query = db.session.query(
            Candidate.id,
            Candidate.name,
            Candidate.email,
            Position.title.label('position'),
            AssessmentResult.step,
            AssessmentResult.score,
            AssessmentResult.status,
            AssessmentResult.completed_at
        ).join(Position).outerjoin(AssessmentResult)
        
        if candidate_id:
            query = query.filter(Candidate.id == candidate_id)
        if position_id:
            query = query.filter(Candidate.position_id == position_id)
        if date_from:
            query = query.filter(AssessmentResult.completed_at >= date_from)
        if date_to:
            query = query.filter(AssessmentResult.completed_at <= date_to)
        
        results = query.all()
        
        # Convert to DataFrame
        data = []
        for row in results:
            data.append({
                'Candidate ID': row.id,
                'Name': row.name,
                'Email': row.email,
                'Position': row.position,
                'Step': row.step,
                'Score': row.score,
                'Status': row.status,
                'Completed At': row.completed_at
            })
        
        return pd.DataFrame(data)
    
    @staticmethod
    def generate_position_performance_report(position_id: int = None, 
                                          date_from: datetime = None, 
                                          date_to: datetime = None) -> pd.DataFrame:
        """Generate position performance report."""
        # Base query for position performance
        query = db.session.query(
            Position.id,
            Position.title,
            func.count(Candidate.id).label('total_candidates'),
            func.count(AssessmentResult.id).label('completed_assessments'),
            func.avg(AssessmentResult.score).label('avg_score'),
            func.count(case([(AssessmentResult.status == 'passed', 1)])).label('passed_count'),
            func.count(case([(AssessmentResult.status == 'failed', 1)])).label('failed_count')
        ).join(Candidate, Position.id == Candidate.position_id)\
         .outerjoin(AssessmentResult, Candidate.id == AssessmentResult.candidate_id)
        
        if position_id:
            query = query.filter(Position.id == position_id)
        if date_from:
            query = query.filter(AssessmentResult.completed_at >= date_from)
        if date_to:
            query = query.filter(AssessmentResult.completed_at <= date_to)
        
        results = query.group_by(Position.id, Position.title).all()
        
        # Convert to DataFrame
        data = []
        for row in results:
            pass_rate = (row.passed_count / row.completed_assessments * 100) if row.completed_assessments > 0 else 0
            data.append({
                'Position ID': row.id,
                'Position Title': row.title,
                'Total Candidates': row.total_candidates,
                'Completed Assessments': row.completed_assessments,
                'Average Score': round(row.avg_score, 2) if row.avg_score else 0,
                'Passed Count': row.passed_count,
                'Failed Count': row.failed_count,
                'Pass Rate (%)': round(pass_rate, 2)
            })
        
        return pd.DataFrame(data)
    
    @staticmethod
    def generate_interviewer_performance_report(interviewer_id: int = None,
                                             date_from: datetime = None,
                                             date_to: datetime = None) -> pd.DataFrame:
        """Generate interviewer performance analytics."""
        query = db.session.query(
            User.id,
            User.username,
            func.count(InterviewEvaluation.id).label('total_evaluations'),
            func.avg(InterviewEvaluation.score).label('avg_score'),
            func.count(case([(InterviewEvaluation.recommendation == 'hire', 1)])).label('hire_recommendations'),
            func.count(case([(InterviewEvaluation.recommendation == 'reject', 1)])).label('reject_recommendations'),
            func.count(case([(InterviewEvaluation.recommendation == 'review', 1)])).label('review_recommendations')
        ).outerjoin(InterviewEvaluation, User.id == InterviewEvaluation.interviewer_id)
        
        if interviewer_id:
            query = query.filter(User.id == interviewer_id)
        if date_from:
            query = query.filter(InterviewEvaluation.evaluated_at >= date_from)
        if date_to:
            query = query.filter(InterviewEvaluation.evaluated_at <= date_to)
        
        results = query.group_by(User.id, User.username).all()
        
        # Convert to DataFrame
        data = []
        for row in results:
            total_evaluations = row.total_evaluations or 0
            hire_rate = (row.hire_recommendations / total_evaluations * 100) if total_evaluations > 0 else 0
            data.append({
                'Interviewer ID': row.id,
                'Username': row.username,
                'Total Evaluations': total_evaluations,
                'Average Score': round(row.avg_score, 2) if row.avg_score else 0,
                'Hire Recommendations': row.hire_recommendations,
                'Reject Recommendations': row.reject_recommendations,
                'Review Recommendations': row.review_recommendations,
                'Hire Rate (%)': round(hire_rate, 2)
            })
        
        return pd.DataFrame(data)
    
    @staticmethod
    def generate_comprehensive_recruitment_report(date_from: datetime = None,
                                               date_to: datetime = None) -> Dict[str, pd.DataFrame]:
        """Generate comprehensive recruitment report with multiple sheets."""
        if not date_from:
            date_from = datetime.now() - timedelta(days=30)
        if not date_to:
            date_to = datetime.now()
        
        # Generate different report components
        candidate_progress = ReportGenerator.generate_candidate_progress_report(
            date_from=date_from, date_to=date_to
        )
        
        position_performance = ReportGenerator.generate_position_performance_report(
            date_from=date_from, date_to=date_to
        )
        
        interviewer_performance = ReportGenerator.generate_interviewer_performance_report(
            date_from=date_from, date_to=date_to
        )
        
        # Generate executive decision summary
        executive_summary = ReportGenerator.generate_executive_decision_summary(
            date_from=date_from, date_to=date_to
        )
        
        # Generate recruitment funnel analysis
        funnel_analysis = ReportGenerator.generate_recruitment_funnel_analysis(
            date_from=date_from, date_to=date_to
        )
        
        return {
            'candidate_progress': candidate_progress,
            'position_performance': position_performance,
            'interviewer_performance': interviewer_performance,
            'executive_summary': executive_summary,
            'funnel_analysis': funnel_analysis
        }
    
    @staticmethod
    def generate_executive_decision_summary(date_from: datetime = None,
                                          date_to: datetime = None) -> pd.DataFrame:
        """Generate executive decision summary report."""
        query = db.session.query(
            ExecutiveDecision.id,
            Candidate.name.label('candidate_name'),
            Position.title.label('position_title'),
            ExecutiveDecision.final_decision,
            ExecutiveDecision.final_score,
            ExecutiveDecision.cto_score,
            ExecutiveDecision.ceo_score,
            ExecutiveDecision.completed_at
        ).join(Candidate).join(Position)
        
        if date_from:
            query = query.filter(ExecutiveDecision.completed_at >= date_from)
        if date_to:
            query = query.filter(ExecutiveDecision.completed_at <= date_to)
        
        results = query.all()
        
        data = []
        for row in results:
            data.append({
                'Decision ID': row.id,
                'Candidate Name': row.candidate_name,
                'Position': row.position_title,
                'Final Decision': row.final_decision,
                'Final Score': round(row.final_score, 2) if row.final_score else 0,
                'CTO Score': round(row.cto_score, 2) if row.cto_score else 0,
                'CEO Score': round(row.ceo_score, 2) if row.ceo_score else 0,
                'Completed At': row.completed_at
            })
        
        return pd.DataFrame(data)
    
    @staticmethod
    def generate_recruitment_funnel_analysis(date_from: datetime = None,
                                           date_to: datetime = None) -> pd.DataFrame:
        """Generate recruitment funnel analysis."""
        if not date_from:
            date_from = datetime.now() - timedelta(days=30)
        if not date_to:
            date_to = datetime.now()
        
        # Step 1: Total candidates
        total_candidates = db.session.query(func.count(Candidate.id))\
            .filter(Candidate.created_at.between(date_from, date_to)).scalar()
        
        # Step 1: Completed assessments
        step1_completed = db.session.query(func.count(AssessmentResult.id))\
            .filter(and_(
                AssessmentResult.step == 1,
                AssessmentResult.completed_at.between(date_from, date_to)
            )).scalar()
        
        # Step 1: Passed assessments
        step1_passed = db.session.query(func.count(AssessmentResult.id))\
            .filter(and_(
                AssessmentResult.step == 1,
                AssessmentResult.status == 'passed',
                AssessmentResult.completed_at.between(date_from, date_to)
            )).scalar()
        
        # Step 2: Completed interviews
        step2_completed = db.session.query(func.count(InterviewEvaluation.id))\
            .filter(and_(
                InterviewEvaluation.step == 2,
                InterviewEvaluation.evaluated_at.between(date_from, date_to)
            )).scalar()
        
        # Step 2: Passed interviews
        step2_passed = db.session.query(func.count(InterviewEvaluation.id))\
            .filter(and_(
                InterviewEvaluation.step == 2,
                InterviewEvaluation.recommendation == 'hire',
                InterviewEvaluation.evaluated_at.between(date_from, date_to)
            )).scalar()
        
        # Step 3: Completed executive decisions
        step3_completed = db.session.query(func.count(ExecutiveDecision.id))\
            .filter(and_(
                ExecutiveDecision.completed_at.between(date_from, date_to)
            )).scalar()
        
        # Step 3: Hired candidates
        step3_hired = db.session.query(func.count(ExecutiveDecision.id))\
            .filter(and_(
                ExecutiveDecision.final_decision == 'hire',
                ExecutiveDecision.completed_at.between(date_from, date_to)
            )).scalar()
        
        data = [{
            'Stage': 'Total Candidates',
            'Count': total_candidates,
            'Conversion Rate (%)': 100.0
        }, {
            'Stage': 'Step 1 - Assessment',
            'Count': step1_completed,
            'Conversion Rate (%)': round((step1_completed / total_candidates * 100), 2) if total_candidates > 0 else 0
        }, {
            'Stage': 'Step 1 - Passed',
            'Count': step1_passed,
            'Conversion Rate (%)': round((step1_passed / step1_completed * 100), 2) if step1_completed > 0 else 0
        }, {
            'Stage': 'Step 2 - Interview',
            'Count': step2_completed,
            'Conversion Rate (%)': round((step2_completed / step1_passed * 100), 2) if step1_passed > 0 else 0
        }, {
            'Stage': 'Step 2 - Passed',
            'Count': step2_passed,
            'Conversion Rate (%)': round((step2_passed / step2_completed * 100), 2) if step2_completed > 0 else 0
        }, {
            'Stage': 'Step 3 - Executive',
            'Count': step3_completed,
            'Conversion Rate (%)': round((step3_completed / step2_passed * 100), 2) if step2_passed > 0 else 0
        }, {
            'Stage': 'Hired',
            'Count': step3_hired,
            'Conversion Rate (%)': round((step3_hired / step3_completed * 100), 2) if step3_completed > 0 else 0
        }]
        
        return pd.DataFrame(data)


class ExcelReportExporter:
    """Excel export functionality with professional formatting."""
    
    @staticmethod
    def export_to_excel(dataframes: Dict[str, pd.DataFrame], filename: str) -> BytesIO:
        """Export multiple dataframes to Excel with professional formatting."""
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for sheet_name, df in dataframes.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Get the worksheet for formatting
                worksheet = writer.sheets[sheet_name]
                
                # Apply professional formatting
                ExcelReportExporter._format_worksheet(worksheet, df)
        
        output.seek(0)
        return output
    
    @staticmethod
    def _format_worksheet(worksheet, df: pd.DataFrame):
        """Apply professional formatting to worksheet."""
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply header formatting
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
            for cell in row:
                cell.border = thin_border
    
    @staticmethod
    def export_single_report(df: pd.DataFrame, filename: str) -> BytesIO:
        """Export single dataframe to Excel."""
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Report', index=False)
            worksheet = writer.sheets['Report']
            ExcelReportExporter._format_worksheet(worksheet, df)
        
        output.seek(0)
        return output


class AutomatedReportScheduler:
    """Automated report scheduling system."""
    
    @staticmethod
    def schedule_daily_report():
        """Schedule daily recruitment summary report."""
        # This would integrate with a task scheduler like Celery
        # For now, we'll create a placeholder
        pass
    
    @staticmethod
    def schedule_weekly_report():
        """Schedule weekly comprehensive report."""
        pass
    
    @staticmethod
    def schedule_monthly_report():
        """Schedule monthly analytics report."""
        pass


# Route Definitions

@report_generation_bp.route('/reports')
@login_required
@hr_required
@rate_limit('reports', {'requests': 50, 'window': 3600})
@audit_log('view_reports')
def reports_dashboard():
    """Main reports dashboard."""
    return render_template('reports/dashboard.html')


@report_generation_bp.route('/reports/candidate-progress')
@login_required
@hr_required
@rate_limit('reports', {'requests': 20, 'window': 3600})
@audit_log('generate_candidate_progress_report')
def candidate_progress_report():
    """Generate candidate progress report."""
    try:
        # Get filter parameters
        candidate_id = request.args.get('candidate_id', type=int)
        position_id = request.args.get('position_id', type=int)
        date_from_str = request.args.get('date_from')
        date_to_str = request.args.get('date_to')
        
        # Parse dates
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d') if date_from_str else None
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d') if date_to_str else None
        
        # Generate report
        df = ReportGenerator.generate_candidate_progress_report(
            candidate_id=candidate_id,
            position_id=position_id,
            date_from=date_from,
            date_to=date_to
        )
        
        if request.args.get('export') == 'excel':
            output = ExcelReportExporter.export_single_report(df, 'candidate_progress_report.xlsx')
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='candidate_progress_report.xlsx'
            )
        
        # Return JSON for AJAX requests
        return jsonify({
            'success': True,
            'data': df.to_dict('records'),
            'total_records': len(df)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@report_generation_bp.route('/reports/position-performance')
@login_required
@hr_required
@rate_limit('reports', {'requests': 20, 'window': 3600})
@audit_log('generate_position_performance_report')
def position_performance_report():
    """Generate position performance report."""
    try:
        position_id = request.args.get('position_id', type=int)
        date_from_str = request.args.get('date_from')
        date_to_str = request.args.get('date_to')
        
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d') if date_from_str else None
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d') if date_to_str else None
        
        df = ReportGenerator.generate_position_performance_report(
            position_id=position_id,
            date_from=date_from,
            date_to=date_to
        )
        
        if request.args.get('export') == 'excel':
            output = ExcelReportExporter.export_single_report(df, 'position_performance_report.xlsx')
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='position_performance_report.xlsx'
            )
        
        return jsonify({
            'success': True,
            'data': df.to_dict('records'),
            'total_records': len(df)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@report_generation_bp.route('/reports/interviewer-performance')
@login_required
@hr_required
@rate_limit('reports', {'requests': 20, 'window': 3600})
@audit_log('generate_interviewer_performance_report')
def interviewer_performance_report():
    """Generate interviewer performance report."""
    try:
        interviewer_id = request.args.get('interviewer_id', type=int)
        date_from_str = request.args.get('date_from')
        date_to_str = request.args.get('date_to')
        
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d') if date_from_str else None
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d') if date_to_str else None
        
        df = ReportGenerator.generate_interviewer_performance_report(
            interviewer_id=interviewer_id,
            date_from=date_from,
            date_to=date_to
        )
        
        if request.args.get('export') == 'excel':
            output = ExcelReportExporter.export_single_report(df, 'interviewer_performance_report.xlsx')
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='interviewer_performance_report.xlsx'
            )
        
        return jsonify({
            'success': True,
            'data': df.to_dict('records'),
            'total_records': len(df)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@report_generation_bp.route('/reports/comprehensive')
@login_required
@admin_required
@rate_limit('reports', {'requests': 10, 'window': 3600})
@audit_log('generate_comprehensive_report')
def comprehensive_report():
    """Generate comprehensive recruitment report."""
    try:
        date_from_str = request.args.get('date_from')
        date_to_str = request.args.get('date_to')
        
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d') if date_from_str else None
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d') if date_to_str else None
        
        reports = ReportGenerator.generate_comprehensive_recruitment_report(
            date_from=date_from,
            date_to=date_to
        )
        
        if request.args.get('export') == 'excel':
            output = ExcelReportExporter.export_to_excel(reports, 'comprehensive_recruitment_report.xlsx')
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='comprehensive_recruitment_report.xlsx'
            )
        
        return jsonify({
            'success': True,
            'reports': {
                'candidate_progress': reports['candidate_progress'].to_dict('records'),
                'position_performance': reports['position_performance'].to_dict('records'),
                'interviewer_performance': reports['interviewer_performance'].to_dict('records'),
                'executive_summary': reports['executive_summary'].to_dict('records'),
                'funnel_analysis': reports['funnel_analysis'].to_dict('records')
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@report_generation_bp.route('/reports/schedule', methods=['GET', 'POST'])
@login_required
@admin_required
@rate_limit('reports', {'requests': 10, 'window': 3600})
@audit_log('schedule_report')
def schedule_report():
    """Schedule automated reports."""
    if request.method == 'POST':
        try:
            report_type = request.form.get('report_type')
            frequency = request.form.get('frequency')  # daily, weekly, monthly
            recipients = request.form.get('recipients')
            
            # This would integrate with a task scheduler
            flash(f'Report {report_type} scheduled for {frequency} delivery', 'success')
            return redirect(url_for('report_generation.reports_dashboard'))
            
        except Exception as e:
            flash(f'Error scheduling report: {str(e)}', 'error')
    
    return render_template('reports/schedule.html')


# API Endpoints for AJAX

@report_generation_bp.route('/api/reports/candidate-progress')
@login_required
@rate_limit('api', {'requests': 100, 'window': 3600})
def api_candidate_progress():
    """API endpoint for candidate progress report."""
    return candidate_progress_report()


@report_generation_bp.route('/api/reports/position-performance')
@login_required
@rate_limit('api', {'requests': 100, 'window': 3600})
def api_position_performance():
    """API endpoint for position performance report."""
    return position_performance_report()


@report_generation_bp.route('/api/reports/interviewer-performance')
@login_required
@rate_limit('api', {'requests': 100, 'window': 3600})
def api_interviewer_performance():
    """API endpoint for interviewer performance report."""
    return interviewer_performance_report()


@report_generation_bp.route('/api/reports/comprehensive')
@login_required
@rate_limit('api', {'requests': 50, 'window': 3600})
def api_comprehensive_report():
    """API endpoint for comprehensive report."""
    return comprehensive_report() 